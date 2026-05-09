from pathlib import Path
import openpyxl
from scripts.init_workbook import create_workbook


def test_create_workbook_has_four_named_sheets(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Dashboard", "Movimientos", "Categorías", "_aux"]


def test_create_workbook_fails_if_file_exists(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    import pytest
    with pytest.raises(FileExistsError):
        create_workbook(out)


def test_movimientos_has_table_with_expected_columns(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Movimientos"]
    assert "tblMov" in ws.tables
    headers = [c.value for c in ws[1]]
    assert headers == ["Fecha", "Tipo", "Categoría", "Importe", "Descripción"]


def test_categorias_sheet_lists_all_categories(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Categorías"]
    pairs = {(row[0].value, row[1].value) for row in ws.iter_rows(min_row=2) if row[0].value}
    # Spot-check: at least these pairs must be present
    assert ("Gasto", "Alimentación") in pairs
    assert ("Gasto", "Restauración") in pairs
    assert ("Ingreso", "Nómina") in pairs
    # And total count matches our source of truth
    from scripts.categories import CATEGORIES
    expected = sum(len(v) for v in CATEGORIES.values())
    assert len(pairs) == expected


def test_movimientos_has_data_validations(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Movimientos"]
    # At least two validations attached: one for Tipo (col B), one for Categoría (col C)
    dv_ranges = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert any("B" in r for r in dv_ranges), f"Missing Tipo validation, got {dv_ranges}"
    assert any("C" in r for r in dv_ranges), f"Missing Categoría validation, got {dv_ranges}"


def test_aux_has_selected_month_and_kpi_formulas(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]

    # v2.0: Selected month is driven from Dashboard!P1
    assert aux["B1"].value == "=Dashboard!P1"

    # KPIs use SUMIFS over tblMov filtered by month
    ingresos = aux["D1"].value or ""
    gastos = aux["D2"].value or ""
    assert "SUMIFS" in ingresos.upper()
    assert "tblMov" in ingresos
    assert "SUMIFS" in gastos.upper()

    # Ahorro is Ingresos - Gastos (or arithmetic of D1/D2)
    ahorro = aux["D3"].value or ""
    assert "D1" in ahorro and "D2" in ahorro

    # % ahorro guards against /0
    pct = aux["D4"].value or ""
    assert "SI.ERROR" in pct.upper() or "IFERROR" in pct.upper()


def test_aux_category_block_has_ten_rows(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]

    cats = [aux.cell(row=r, column=6).value for r in range(2, 12)]
    sums = [aux.cell(row=r, column=7).value for r in range(2, 12)]
    assert all(cats), f"Missing category labels: {cats}"
    assert all(s and "SUMIFS" in s.upper() for s in sums), f"Missing SUMIFS: {sums}"


def test_aux_monthly_evolution_block_has_twelve_rows(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]

    # Row 2 is January, row 13 is December
    months = [aux.cell(row=r, column=9).value for r in range(2, 14)]
    ingresos = [aux.cell(row=r, column=10).value for r in range(2, 14)]
    gastos = [aux.cell(row=r, column=11).value for r in range(2, 14)]
    assert all(m for m in months), f"Missing month labels: {months}"
    assert all("SUMIFS" in (s or "").upper() for s in ingresos)
    assert all("SUMIFS" in (s or "").upper() for s in gastos)


def test_dashboard_has_month_selector_default_to_today(tmp_path: Path):
    """v2.0: month selector is at P1 (not B1)."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]

    formula = (dash["P1"].value or "").upper()
    assert "FECHA" in formula or "DATE" in formula
    assert "HOY" in formula or "TODAY" in formula


def test_dashboard_kpis_use_named_ranges(tmp_path: Path):
    """v2.0: KPI value cells use named range references."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]

    assert dash["B7"].value == "=KPI_Ingresos"
    assert dash["G7"].value == "=KPI_Gastos"
    assert dash["L7"].value == "=KPI_Ahorro"
    assert dash["Q7"].value == "=KPI_PctAhorro"


def test_dashboard_top5_links_to_aux(tmp_path: Path):
    """v2.0: Top-5 is at K14:S18."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]

    # 5 rows starting at row 14: K=date, M=description, S=amount
    for i in range(5):
        data_row = 14 + i
        aux_row = 2 + i
        assert dash.cell(row=data_row, column=11).value == f"=_aux!N{aux_row}", \
            f"K{data_row} should be =_aux!N{aux_row}"
        assert dash.cell(row=data_row, column=13).value == f"=_aux!O{aux_row}", \
            f"M{data_row} should be =_aux!O{aux_row}"
        assert dash.cell(row=data_row, column=19).value == f"=_aux!P{aux_row}", \
            f"S{data_row} should be =_aux!P{aux_row}"


def test_dashboard_has_pie_and_line_charts(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]

    chart_types = {type(c).__name__ for c in dash._charts}
    assert "PieChart" in chart_types, f"got {chart_types}"
    assert "LineChart" in chart_types, f"got {chart_types}"


def test_aux_sheet_is_hidden(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)

    wb = openpyxl.load_workbook(out)
    assert wb["_aux"].sheet_state == "hidden"
    # Dashboard is the active/visible default sheet
    assert wb.active.title == "Dashboard"


def test_aux_evolution_is_calendar_year(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]
    # Row 2 = January of YEAR($B$1); row 13 = December of YEAR($B$1)
    row2 = aux.cell(row=2, column=9).value or ""
    row13 = aux.cell(row=13, column=9).value or ""
    assert "DATE(YEAR($B$1),1,1)" in row2, f"row 2 should be January: {row2}"
    assert "DATE(YEAR($B$1),12,1)" in row13, f"row 13 should be December: {row13}"


def test_aux_evolution_has_pct_ahorro_column(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]
    assert aux["M1"].value == "% Ahorro"
    for r in range(2, 14):
        f = aux.cell(row=r, column=13).value or ""
        assert "IFERROR" in f.upper() and f"L{r}" in f and f"J{r}" in f


def test_aux_top5_uses_helper_column(tmp_path: Path):
    """v2.0: Top-5 uses LARGE + helper column Q (no AGGREGATE)."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]

    # P2 should use LARGE over $Q$2:$Q$1000
    p2 = (aux["P2"].value or "")
    assert "LARGE($Q$" in p2, f"P2 should use LARGE($Q$...): {p2}"

    # N2 should use INDEX/MATCH (not SUMPRODUCT)
    n2 = (aux["N2"].value or "").upper()
    assert "MATCH" in n2, f"N2 should use MATCH: {n2}"
    assert "AGGREGATE" not in n2, f"N2 should NOT use AGGREGATE: {n2}"


def test_aux_helper_column_q_filters_by_month(tmp_path: Path):
    """v2.0: _aux!Q2 uses per-row IF to filter expenses for the month."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]

    # Q1 has the helper header
    assert aux["Q1"].value == "_helper_top5", f"Q1 header: {aux['Q1'].value}"

    # Q2 formula filters by Tipo="Gasto" and month range
    q2 = (aux["Q2"].value or "")
    assert 'IF(AND(Movimientos!B2="Gasto"' in q2, f"Q2 formula: {q2}"
    assert "MONTH($B$1)+1" in q2, f"Q2 missing month end: {q2}"


def test_aux_cumulative_daily_block(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]
    assert aux["R1"].value == "Día"
    assert aux["S1"].value == "Acumulado"
    # 31 rows of dates (R2..R32) and 31 rows of cumulative SUMIFS (S2..S32)
    for r in range(2, 33):
        date_formula = aux.cell(row=r, column=18).value or ""
        sum_formula = (aux.cell(row=r, column=19).value or "").upper()
        assert "$B$1" in date_formula
        assert "SUMIFS" in sum_formula and "NA()" in sum_formula


def test_aux_category_comparison_block(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    aux = wb["_aux"]
    assert aux["U1"].value == "Categoría"
    assert aux["V1"].value == "Este mes"
    assert aux["W1"].value == "Mes anterior"
    from scripts.categories import CATEGORIES
    cats = [aux.cell(row=r, column=21).value for r in range(2, 12)]
    assert cats == CATEGORIES["Gasto"]
    for r in range(2, 12):
        v = (aux.cell(row=r, column=22).value or "").upper()
        w = (aux.cell(row=r, column=23).value or "").upper()
        assert "SUMIFS" in v and "SUMIFS" in w


def test_dashboard_has_five_charts(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]
    types = [type(c).__name__ for c in dash._charts]
    assert types.count("PieChart") == 1
    assert types.count("LineChart") == 3
    assert types.count("BarChart") == 1
    assert len(types) == 5


def test_dashboard_uses_arial_font(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]
    # v2.0: header at A1 (merged), value cells at B7, G7
    assert dash["A1"].font.name == "Arial"
    assert dash["B7"].font.name == "Arial"


def test_dashboard_kpi_values_are_ink_black(tmp_path: Path):
    """v2.0: KPI value cells use INK (#111111), not green (conditional formatting handles colour)."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]
    # Base font for Ingresos and Gastos cards should be INK
    for cell_ref in ("B7", "G7"):
        color = (dash[cell_ref].font.color.rgb or "").upper()
        assert "111111" in color, f"{cell_ref} not ink: {color}"


def test_month_selector_is_blue_input(tmp_path: Path):
    """v2.0: month selector moved to P1."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]
    color = (dash["P1"].font.color.rgb or "").upper()
    assert "0000FF" in color, f"P1 not blue: {color}"


def test_currency_format_has_parens_for_negatives(tmp_path: Path):
    """v2.0: check B7 (Ingresos KPI cell) has parenthesis format."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]
    fmt = dash["B7"].number_format
    assert "(" in fmt and ")" in fmt, f"B7 format missing parens: {fmt}"
    assert '"-"' in fmt, f"B7 format missing zero literal: {fmt}"


# ---------------------------------------------------------------------------
# New tests (v2.0)
# ---------------------------------------------------------------------------

def test_dashboard_has_named_ranges(tmp_path: Path):
    """All 5 named ranges must exist in the workbook."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)

    for name in ("Mes_Seleccionado", "KPI_Ingresos", "KPI_Gastos", "KPI_Ahorro", "KPI_PctAhorro"):
        assert name in wb.defined_names, f"Missing named range: {name}"


def test_dashboard_freeze_panes_at_row_11(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]
    assert dash.freeze_panes == "A11", f"freeze_panes: {dash.freeze_panes}"


def test_dashboard_gridlines_hidden(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]
    assert dash.sheet_view.showGridLines == False, \
        f"showGridLines should be False: {dash.sheet_view.showGridLines}"


def test_dashboard_protected_with_selector_unlocked(tmp_path: Path):
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]

    assert dash.protection.sheet == True, "Sheet should be protected"
    assert dash["P1"].protection.locked == False, "P1 (month selector) should be unlocked"


def test_dashboard_has_conditional_formatting_on_ahorro(tmp_path: Path):
    """At least 2 FormulaRules must apply to L7 or Q7 (Ahorro/% Ahorro KPIs)."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]

    formula_rule_count = 0
    for cf_range, rules in dash.conditional_formatting._cf_rules.items():
        range_str = str(cf_range)
        if "L7" in range_str or "Q7" in range_str:
            for rule in rules:
                if rule.type == "expression":
                    formula_rule_count += 1

    assert formula_rule_count >= 2, \
        f"Expected >=2 FormulaRules on L7/Q7, got {formula_rule_count}"


def test_dashboard_top5_has_data_bar_rule(tmp_path: Path):
    """A DataBarRule must be registered covering S14:S18."""
    out = tmp_path / "gastos.xlsx"
    create_workbook(out)
    wb = openpyxl.load_workbook(out)
    dash = wb["Dashboard"]

    found = False
    for cf_range, rules in dash.conditional_formatting._cf_rules.items():
        range_str = str(cf_range)
        if "S14" in range_str or "S18" in range_str:
            for rule in rules:
                if rule.type == "dataBar":
                    found = True
                    break

    assert found, "No DataBarRule found covering S14:S18"
