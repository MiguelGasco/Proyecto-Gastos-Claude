"""Verify gastos.xlsx has zero formula errors after LibreOffice recalc.

Usage:
    python -m scripts.verify_workbook [path]

Default path: gastos.xlsx in cwd. Exits 0 if clean, 1 if any cell contains
an Excel error marker after LibreOffice has recalculated.
"""
from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import List, Tuple

import openpyxl


WINDOWS_LO_PATHS = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]
ERROR_MARKERS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#N/A", "#NUM!")


class LibreOfficeNotFoundError(RuntimeError):
    pass


def find_soffice() -> str:
    """Return the path to soffice.exe / soffice on this system."""
    if sys.platform == "win32":
        for p in WINDOWS_LO_PATHS:
            if os.path.exists(p):
                return p
        # Fall back to PATH
        if shutil.which("soffice"):
            return "soffice"
        raise LibreOfficeNotFoundError(
            "LibreOffice no encontrado. Instálalo desde https://www.libreoffice.org/"
            f" o añade soffice al PATH. Buscado en: {WINDOWS_LO_PATHS}"
        )
    # Linux/macOS
    if shutil.which("soffice"):
        return "soffice"
    raise LibreOfficeNotFoundError("soffice no encontrado en PATH")


def recalc_in_place(xlsx: Path, timeout: int = 60) -> None:
    """Open xlsx in LibreOffice headless, recalculate formulas, save back to same path."""
    soffice = find_soffice()
    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td)
        result = subprocess.run(
            [
                soffice,
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(td_path),
                str(xlsx),
            ],
            capture_output=True,
            timeout=timeout,
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"soffice falló (exit {result.returncode}): "
                f"{result.stderr.decode(errors='ignore')}"
            )

        converted = list(td_path.glob("*.xlsx"))
        if not converted:
            raise RuntimeError(
                f"soffice no produjo .xlsx en {td_path}. stdout: "
                f"{result.stdout.decode(errors='ignore')}"
            )
        shutil.copy2(converted[0], xlsx)


def scan_errors(xlsx: Path) -> List[str]:
    """Return list of 'Sheet!Cell=#ERROR' strings for any error cells."""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    errors: List[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v in ERROR_MARKERS:
                    errors.append(f"{sheet_name}!{cell.coordinate}={v}")
    return errors


def verify(xlsx: Path) -> Tuple[bool, List[str]]:
    recalc_in_place(xlsx)
    errors = scan_errors(xlsx)
    return (len(errors) == 0, errors)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("path", nargs="?", default="gastos.xlsx")
    args = parser.parse_args()

    xlsx = Path(args.path).resolve()
    if not xlsx.exists():
        print(f"ERROR: no existe {xlsx}", file=sys.stderr)
        return 2

    try:
        ok, errors = verify(xlsx)
    except LibreOfficeNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 3
    except Exception as e:
        print(f"ERROR durante recalc: {e}", file=sys.stderr)
        return 4

    if ok:
        print(f"OK: {xlsx.name} sin errores de fórmula.")
        return 0
    else:
        print(f"FALLO: {xlsx.name} tiene {len(errors)} errores:")
        for e in errors[:50]:  # cap output
            print(f"  - {e}")
        if len(errors) > 50:
            print(f"  ... y {len(errors) - 50} más")
        return 1


if __name__ == "__main__":
    sys.exit(main())
