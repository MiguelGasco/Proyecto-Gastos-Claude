"""Build gastos.xlsx from scratch."""
from pathlib import Path
import argparse
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.chart import PieChart, LineChart, BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.workbook.defined_name import DefinedName
from scripts.categories import CATEGORIES, VALID_TYPES


SHEET_ORDER = ["Dashboard", "Movimientos", "Categorías", "_aux"]

MOV_HEADERS = ["Fecha", "Tipo", "Categoría", "Importe", "Descripción"]
MOV_LAST_DATA_ROW = 1000  # pre-size the table for data validations and table ref

# --- Skill-convention number formats ---
_FMT_CURRENCY = '#,##0.00 "€";(#,##0.00 "€");"-"'
_FMT_PERCENT  = '0.0%;(0.0%);"-"'
_FMT_DATE     = "dd/mm/yyyy"
_FMT_DAY      = "dd/mm"

# --- Skill-convention colors ---
_COLOR_BLUE  = "FF0000FF"   # hardcoded inputs
_COLOR_BLACK = "FF000000"   # formulas / calculations
_COLOR_GREEN = "FF008000"   # cross-sheet links

# --- Minimalist palette (v2.0) ---
INK      = "111111"   # primary text
MUTED    = "7A7A7A"   # labels, subtitles
RULE     = "E5E5E5"   # thin separator lines
POSITIVE = "1B7F3A"   # green for positive values
NEGATIVE = "B33A3A"   # red for negative values

# 10-tone grayscale palette for pie chart slices (darkest to lightest)
GREY_PALETTE = [
    "111111", "2E2E2E", "4A4A4A", "666666", "808080",
    "999999", "B3B3B3", "C9C9C9", "DCDCDC", "EEEEEE",
]


# ---------------------------------------------------------------------------
# Font / style helpers
# ---------------------------------------------------------------------------

def _apply_font(cell, *, color: str = _COLOR_BLACK, size: int = 11, bold: bool = False) -> None:
    """Apply Arial font with skill color conventions."""
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)


def _blue_input(cell) -> None:
    """Blue = hardcoded input the user will change."""
    _apply_font(cell, color=_COLOR_BLUE)


def _green_link(cell, *, size: int = 11, bold: bool = False) -> None:
    """Green = cross-sheet link within same workbook."""
    _apply_font(cell, color=_COLOR_GREEN, size=size, bold=bold)


def _black_formula(cell, *, size: int = 11, bold: bool = False) -> None:
    """Black = formula / calculation."""
    _apply_font(cell, color=_COLOR_BLACK, size=size, bold=bold)


def _ink(cell, *, size: int = 11, bold: bool = False) -> None:
    """INK (#111111) = primary text in minimalist palette."""
    cell.font = Font(name="Arial", size=size, bold=bold, color=INK)


def _muted(cell, *, size: int = 11, bold: bool = False) -> None:
    """MUTED (#7A7A7A) = secondary labels in minimalist palette."""
    cell.font = Font(name="Arial", size=size, bold=bold, color=MUTED)


def _currency(cell) -> None:
    cell.number_format = _FMT_CURRENCY


def _percent(cell) -> None:
    cell.number_format = _FMT_PERCENT


# ---------------------------------------------------------------------------
# Chart styling helpers (minimalist)
# ---------------------------------------------------------------------------

def _set_line_color(series, hex_color, dash=False, width_emu=20000):
    """Apply a solid or dashed line color to a chart series."""
    line = LineProperties(solidFill=hex_color, w=width_emu)
    if dash:
        line.prstDash = "dash"
    gp = GraphicalProperties()
    gp.line = line
    series.spPr = gp


def _set_solid_fill(series, hex_color):
    """Apply a solid fill color to a chart series (bar chart)."""
    series.spPr = GraphicalProperties(solidFill=hex_color)


def _apply_pie_grayscale(pie):
    """Apply grayscale palette to pie chart slices."""
    if not pie.series:
        return
    series = pie.series[0]
    series.dPt = [
        DataPoint(idx=i, spPr=GraphicalProperties(solidFill=hex_color))
        for i, hex_color in enumerate(GREY_PALETTE)
    ]


def _minimal_chart_style(chart):
    """Remove major gridlines, move legend to bottom."""
    if hasattr(chart, "x_axis") and chart.x_axis is not None:
        chart.x_axis.majorGridlines = None
    if hasattr(chart, "y_axis") and chart.y_axis is not None:
        chart.y_axis.majorGridlines = None
    if chart.legend is not None:
        chart.legend.position = "b"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_movimientos(ws) -> None:
    for col_idx, header in enumerate(MOV_HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        _black_formula(c, bold=True)

    last_col = get_column_letter(len(MOV_HEADERS))
    table = Table(
        displayName="tblMov",
        ref=f"A1:{last_col}{MOV_LAST_DATA_ROW}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
    )
    ws.add_table(table)

    # Number/date formats on data rows (font/color applied per row in add_movement.py)
    for row in range(2, MOV_LAST_DATA_ROW + 1):
        ws.cell(row=row, column=1).number_format = _FMT_DATE
        ws.cell(row=row, column=4).number_format = _FMT_CURRENCY

    # Sensible column widths
    widths = {"A": 12, "B": 10, "C": 16, "D": 14, "E": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _build_categorias(ws) -> None:
    for col_idx, header in enumerate(["Tipo", "Categoría"], start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        _black_formula(c, bold=True)
    r = 2
    for tipo, cats in CATEGORIES.items():
        for cat in cats:
            c1 = ws.cell(row=r, column=1, value=tipo)
            c2 = ws.cell(row=r, column=2, value=cat)
            _black_formula(c1)
            _black_formula(c2)
            r += 1
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18


def _add_validations(ws_mov) -> None:
    # Tipo: list of two values
    dv_tipo = DataValidation(
        type="list",
        formula1=f'"{",".join(VALID_TYPES)}"',
        allow_blank=True,
    )
    dv_tipo.add(f"B2:B{MOV_LAST_DATA_ROW}")
    ws_mov.add_data_validation(dv_tipo)

    # Categoría: union of all categories
    all_cats = sorted({c for cats in CATEGORIES.values() for c in cats})
    dv_cat = DataValidation(
        type="list",
        formula1=f'"{",".join(all_cats)}"',
        allow_blank=True,
    )
    dv_cat.add(f"C2:C{MOV_LAST_DATA_ROW}")
    ws_mov.add_data_validation(dv_cat)


def _build_aux(ws) -> None:
    # --- B1: selected month (driven by Dashboard!P1) ---
    _black_formula(ws["A1"])
    ws["A1"] = "Mes seleccionado"
    c = ws["B1"]
    c.value = "=Dashboard!P1"          # v2.0: selector moved to P1
    c.number_format = _FMT_DATE
    _black_formula(c)

    _black_formula(ws["A2"])
    ws["A2"] = "Etiqueta"
    c2 = ws["B2"]
    c2.value = '=TEXT(B1,"mm/yyyy")'
    _black_formula(c2)

    # --- KPIs (D1:D4) ---
    month_start = "$B$1"
    month_end_excl = 'DATE(YEAR($B$1),MONTH($B$1)+1,1)'

    ws["C1"] = "Ingresos mes"
    _black_formula(ws["C1"])
    ws["D1"] = (
        f'=SUMIFS(tblMov[Importe],tblMov[Tipo],"Ingreso",'
        f'tblMov[Fecha],">="&{month_start},'
        f'tblMov[Fecha],"<"&{month_end_excl})'
    )
    _black_formula(ws["D1"])
    ws["C2"] = "Gastos mes"
    _black_formula(ws["C2"])
    ws["D2"] = (
        f'=SUMIFS(tblMov[Importe],tblMov[Tipo],"Gasto",'
        f'tblMov[Fecha],">="&{month_start},'
        f'tblMov[Fecha],"<"&{month_end_excl})'
    )
    _black_formula(ws["D2"])
    ws["C3"] = "Ahorro neto"
    _black_formula(ws["C3"])
    ws["D3"] = "=D1-D2"
    _black_formula(ws["D3"])
    ws["C4"] = "% ahorro"
    _black_formula(ws["C4"])
    ws["D4"] = "=IFERROR(D3/D1,0)"
    _black_formula(ws["D4"])

    for r in (1, 2, 3):
        ws.cell(row=r, column=4).number_format = _FMT_CURRENCY
    ws["D4"].number_format = _FMT_PERCENT

    # --- Category-of-month block (F1:G11) ---
    ws["F1"] = "Categoría"
    _black_formula(ws["F1"], bold=True)
    ws["G1"] = "Total mes"
    _black_formula(ws["G1"], bold=True)
    for i, cat in enumerate(CATEGORIES["Gasto"], start=2):
        c_label = ws.cell(row=i, column=6, value=cat)
        _black_formula(c_label)
        c_sum = ws.cell(
            row=i,
            column=7,
            value=(
                f'=SUMIFS(tblMov[Importe],tblMov[Tipo],"Gasto",'
                f'tblMov[Categoría],F{i},'
                f'tblMov[Fecha],">="&{month_start},'
                f'tblMov[Fecha],"<"&{month_end_excl})'
            ),
        )
        _black_formula(c_sum)
        c_sum.number_format = _FMT_CURRENCY

    # --- Monthly evolution block (I1:M13), fixed Jan-Dec of YEAR($B$1) ---
    for hdr, col in [("Mes", 9), ("Ingresos", 10), ("Gastos", 11), ("Ahorro", 12), ("% Ahorro", 13)]:
        c = ws.cell(row=1, column=col, value=hdr)
        _black_formula(c, bold=True)
    for month_num in range(1, 13):
        r = 1 + month_num  # row 2 = January, row 13 = December
        start_ref = f'DATE(YEAR($B$1),{month_num},1)'
        end_ref = f'DATE(YEAR($B$1),{month_num + 1},1)'
        c9 = ws.cell(row=r, column=9, value=f'=TEXT(DATE(YEAR($B$1),{month_num},1),"mm/yyyy")')
        _black_formula(c9)
        c10 = ws.cell(
            row=r,
            column=10,
            value=(
                f'=SUMIFS(tblMov[Importe],tblMov[Tipo],"Ingreso",'
                f'tblMov[Fecha],">="&{start_ref},'
                f'tblMov[Fecha],"<"&{end_ref})'
            ),
        )
        _black_formula(c10)
        c11 = ws.cell(
            row=r,
            column=11,
            value=(
                f'=SUMIFS(tblMov[Importe],tblMov[Tipo],"Gasto",'
                f'tblMov[Fecha],">="&{start_ref},'
                f'tblMov[Fecha],"<"&{end_ref})'
            ),
        )
        _black_formula(c11)
        c12 = ws.cell(row=r, column=12, value=f"=J{r}-K{r}")
        _black_formula(c12)
        c13 = ws.cell(row=r, column=13, value=f"=IFERROR(L{r}/J{r},0)")
        _black_formula(c13)
        for c in (c10, c11, c12):
            c.number_format = _FMT_CURRENCY
        c13.number_format = _FMT_PERCENT

    # --- Top-5 block (N1:P6) with helper column Q ---
    # Header row
    for hdr, col in [("Fecha", 14), ("Descripción", 15), ("Importe", 16)]:
        c = ws.cell(row=1, column=col, value=hdr)
        _black_formula(c, bold=True)

    # Q1: helper column header
    ws.cell(row=1, column=17, value="_helper_top5")
    _black_formula(ws.cell(row=1, column=17))

    # Q2:Q1000 — per-row IF filter (no array operations, compatible with all Excel >= 2010)
    for i in range(2, 1001):
        ws.cell(row=i, column=17, value=(
            f'=IF(AND(Movimientos!B{i}="Gasto",'
            f'Movimientos!A{i}>=$B$1,'
            f'Movimientos!A{i}<DATE(YEAR($B$1),MONTH($B$1)+1,1)),'
            f'Movimientos!D{i},"")'
        ))

    # N2:P6 — LARGE + INDEX/MATCH (classic, no AGGREGATE, no CSE)
    for i in range(5):
        rank = i + 1
        r = 2 + i
        # P column: rank-th largest expense amount
        ws.cell(row=r, column=16, value=f'=IFERROR(LARGE($Q$2:$Q$1000,{rank}),"")')
        ws.cell(row=r, column=16).number_format = _FMT_CURRENCY
        _black_formula(ws.cell(row=r, column=16))
        # N column: date of that expense
        ws.cell(row=r, column=14, value=f'=IFERROR(INDEX(Movimientos!$A$2:$A$1000,MATCH(P{r},$Q$2:$Q$1000,0)),"")')
        ws.cell(row=r, column=14).number_format = _FMT_DATE
        _black_formula(ws.cell(row=r, column=14))
        # O column: description of that expense
        ws.cell(row=r, column=15, value=f'=IFERROR(INDEX(Movimientos!$E$2:$E$1000,MATCH(P{r},$Q$2:$Q$1000,0)),"")')
        _black_formula(ws.cell(row=r, column=15))

    # --- Cumulative daily expenses block (R1:S32) ---
    c = ws["R1"]
    c.value = "Día"
    _black_formula(c, bold=True)
    c = ws["S1"]
    c.value = "Acumulado"
    _black_formula(c, bold=True)
    for i in range(31):
        r = 2 + i
        cr = ws.cell(row=r, column=18, value=f"=$B$1+{i}")
        _black_formula(cr)
        cr.number_format = _FMT_DAY
        cs = ws.cell(
            row=r,
            column=19,
            value=(
                f'=IF(R{r}>=DATE(YEAR($B$1),MONTH($B$1)+1,1),NA(),'
                f'SUMIFS(tblMov[Importe],tblMov[Tipo],"Gasto",'
                f'tblMov[Fecha],">="&$B$1,'
                f'tblMov[Fecha],"<="&R{r}))'
            ),
        )
        _black_formula(cs)
        cs.number_format = _FMT_CURRENCY

    # --- Category comparison: this month vs previous month (U1:W11) ---
    for hdr, col in [("Categoría", 21), ("Este mes", 22), ("Mes anterior", 23)]:
        c = ws.cell(row=1, column=col, value=hdr)
        _black_formula(c, bold=True)
    prev_start = 'DATE(YEAR($B$1),MONTH($B$1)-1,1)'
    for i, cat in enumerate(CATEGORIES["Gasto"], start=2):
        cu = ws.cell(row=i, column=21, value=cat)
        _black_formula(cu)
        cv = ws.cell(
            row=i,
            column=22,
            value=(
                f'=SUMIFS(tblMov[Importe],tblMov[Tipo],"Gasto",'
                f'tblMov[Categoría],U{i},'
                f'tblMov[Fecha],">="&{month_start},'
                f'tblMov[Fecha],"<"&{month_end_excl})'
            ),
        )
        _black_formula(cv)
        cw = ws.cell(
            row=i,
            column=23,
            value=(
                f'=SUMIFS(tblMov[Importe],tblMov[Tipo],"Gasto",'
                f'tblMov[Categoría],U{i},'
                f'tblMov[Fecha],">="&{prev_start},'
                f'tblMov[Fecha],"<"&{month_start})'
            ),
        )
        _black_formula(cw)
        cv.number_format = _FMT_CURRENCY
        cw.number_format = _FMT_CURRENCY

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["O"].width = 30


def _build_dashboard(ws) -> None:
    # -----------------------------------------------------------------------
    # Part 0 — Month selector (top-right, cell P1)
    # -----------------------------------------------------------------------
    o1 = ws["O1"]
    o1.value = "Mes:"
    o1.font = Font(name="Arial", size=10, color=MUTED)
    o1.alignment = Alignment(horizontal="right", vertical="center")

    p1 = ws["P1"]
    p1.value = "=DATE(YEAR(TODAY()),MONTH(TODAY()),1)"
    p1.number_format = _FMT_DATE
    p1.fill = PatternFill("solid", fgColor="FFF2CC")  # soft yellow
    _blue_input(p1)

    # -----------------------------------------------------------------------
    # Part 1 — Header rows 1-3
    # -----------------------------------------------------------------------
    ws.merge_cells("A1:N1")
    a1 = ws["A1"]
    a1.value = "CONTROL DE GASTOS"
    a1.font = Font(name="Arial", size=22, bold=False, color=INK)
    a1.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:N2")
    a2 = ws["A2"]
    a2.value = '=TEXT(P1,"mmmm yyyy")'
    a2.font = Font(name="Arial", size=13, color=MUTED)
    a2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # Row 3: thin separator
    thin_rule = Border(bottom=Side(border_style="thin", color=RULE))
    for col in range(2, 21):  # B to T
        ws.cell(row=3, column=col).border = thin_rule
    ws.row_dimensions[3].height = 6

    # Row 4: spacer
    ws.row_dimensions[4].height = 8

    # -----------------------------------------------------------------------
    # Part 2 — KPI cards (rows 5-9), 4 cards × 4 columns
    # -----------------------------------------------------------------------
    # Cards: B5:E9 | G5:J9 | L5:O9 | Q5:T9
    kpi_cards = [
        ("B", 5, "INGRESOS DEL MES", "=KPI_Ingresos", _FMT_CURRENCY, "B7"),
        ("G", 5, "GASTOS DEL MES",   "=KPI_Gastos",   _FMT_CURRENCY, "G7"),
        ("L", 5, "AHORRO NETO",      "=KPI_Ahorro",   _FMT_CURRENCY, "L7"),
        ("Q", 5, "% AHORRO",         "=KPI_PctAhorro", _FMT_PERCENT, "Q7"),
    ]

    for start_col_letter, start_row, label, formula, fmt, value_cell in kpi_cards:
        # Label row (row 5)
        lbl = ws[f"{start_col_letter}{start_row}"]
        lbl.value = label
        lbl.font = Font(name="Arial", size=9, color=MUTED)
        lbl.alignment = Alignment(horizontal="left", vertical="bottom")

        # Value row (row 7) — large INK number
        vc = ws[value_cell]
        vc.value = formula
        vc.number_format = fmt
        vc.font = Font(name="Arial", size=28, color=INK)
        vc.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 8
    ws.row_dimensions[7].height = 36
    ws.row_dimensions[8].height = 8
    ws.row_dimensions[9].height = 8

    # Conditional formatting: Ahorro (L7) and % Ahorro (Q7)
    for cell_ref in ("L7", "Q7"):
        ws.conditional_formatting.add(
            cell_ref,
            FormulaRule(
                formula=[f"{cell_ref}>=0"],
                font=Font(name="Arial", size=28, color=POSITIVE),
            ),
        )
        ws.conditional_formatting.add(
            cell_ref,
            FormulaRule(
                formula=[f"{cell_ref}<0"],
                font=Font(name="Arial", size=28, color=NEGATIVE),
            ),
        )

    # Row 10: separator line
    sep_border = Border(bottom=Side(border_style="thin", color=RULE))
    for col in range(2, 21):  # B to T
        ws.cell(row=10, column=col).border = sep_border
    ws.row_dimensions[10].height = 12

    # -----------------------------------------------------------------------
    # Part 3 — Section: Resumen del mes (rows 12-29)
    # -----------------------------------------------------------------------
    ws.row_dimensions[11].height = 10

    b12 = ws["B12"]
    b12.value = "Resumen del mes"
    b12.font = Font(name="Arial", size=12, bold=True, color=INK)
    ws.row_dimensions[12].height = 18

    # Top-5 table header row 13 (right side, columns K-T)
    top5_header_border = Border(bottom=Side(border_style="thin", color=RULE))
    for col_letter, label in [("K", "FECHA"), ("M", "DESCRIPCIÓN"), ("S", "IMPORTE")]:
        c = ws[f"{col_letter}13"]
        c.value = label
        c.font = Font(name="Arial", size=10, color=MUTED)
        c.alignment = Alignment(horizontal="left", vertical="bottom")
        c.border = top5_header_border

    ws.row_dimensions[13].height = 16

    # Top-5 data rows 14-18
    for i in range(5):
        data_row = 14 + i
        aux_row = 2 + i

        # K: Date
        k = ws.cell(row=data_row, column=11)  # K
        k.value = f"=_aux!N{aux_row}"
        k.number_format = _FMT_DATE
        _green_link(k)

        # M:R merged for description
        ws.merge_cells(f"M{data_row}:R{data_row}")
        m = ws.cell(row=data_row, column=13)  # M
        m.value = f"=_aux!O{aux_row}"
        _green_link(m)

        # S: Amount
        s = ws.cell(row=data_row, column=19)  # S
        s.value = f"=_aux!P{aux_row}"
        s.number_format = _FMT_CURRENCY
        s.alignment = Alignment(horizontal="right")
        _green_link(s)

        ws.row_dimensions[data_row].height = 16

    # DataBar on S14:S18
    db_rule = DataBarRule(
        start_type="min",
        start_value=0,
        end_type="max",
        end_value=100,
        color=INK,
        showValue=True,
    )
    ws.conditional_formatting.add("S14:S18", db_rule)

    ws.row_dimensions[19].height = 8   # spacer below top-5
    ws.row_dimensions[29].height = 8   # bottom of section

    # -----------------------------------------------------------------------
    # Part 4 — Section: Tendencias 2026 (rows 31-49)
    # -----------------------------------------------------------------------
    ws.row_dimensions[30].height = 10

    b31 = ws["B31"]
    b31.value = "Tendencias 2026"
    b31.font = Font(name="Arial", size=12, bold=True, color=INK)
    ws.row_dimensions[31].height = 18

    # -----------------------------------------------------------------------
    # Part 5 — Section: Análisis del mes (rows 51-69)
    # -----------------------------------------------------------------------
    ws.row_dimensions[50].height = 10

    b51 = ws["B51"]
    b51.value = "Análisis del mes"
    b51.font = Font(name="Arial", size=12, bold=True, color=INK)
    ws.row_dimensions[51].height = 18

    # -----------------------------------------------------------------------
    # Part 6 — Section: Comparativas (rows 71-89)
    # -----------------------------------------------------------------------
    ws.row_dimensions[70].height = 10

    b71 = ws["B71"]
    b71.value = "Categorías: este mes vs anterior"
    b71.font = Font(name="Arial", size=12, bold=True, color=INK)
    ws.row_dimensions[71].height = 18

    # -----------------------------------------------------------------------
    # Column widths
    # -----------------------------------------------------------------------
    ws.column_dimensions["A"].width = 2   # left gutter
    ws.column_dimensions["U"].width = 2   # right gutter
    # P column (month selector) wider
    ws.column_dimensions["P"].width = 16
    # All other columns B-T: width 9
    for col_letter in "BCDEFGHIJKLMNOQRST":
        ws.column_dimensions[col_letter].width = 9


def _add_charts(wb) -> None:
    aux = wb["_aux"]
    dash = wb["Dashboard"]

    # --- (1) Pie chart: gastos por categoría del mes seleccionado ---
    pie = PieChart()
    pie.title = "Gastos por categoría"
    labels = Reference(aux, min_col=6, min_row=2, max_row=11)        # F2:F11
    data = Reference(aux, min_col=7, min_row=1, max_row=11)          # G1:G11 (incl header)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 10   # cm
    pie.width = 12
    _apply_pie_grayscale(pie)
    _minimal_chart_style(pie)
    dash.add_chart(pie, "B14")

    # --- (2) Line chart: evolución mensual (Tendencias 2026 section) ---
    line = LineChart()
    line.title = "Evolución mensual 2026"
    cats = Reference(aux, min_col=9, min_row=2, max_row=13)          # I2:I13
    series_data = Reference(aux, min_col=10, max_col=12, min_row=1, max_row=13)
    line.add_data(series_data, titles_from_data=True)
    line.set_categories(cats)
    line.height = 9
    line.width = 24
    # Style: series 0=Ingresos INK solid, 1=Gastos INK dashed, 2=Ahorro MUTED solid
    if len(line.series) >= 1:
        _set_line_color(line.series[0], INK)
    if len(line.series) >= 2:
        _set_line_color(line.series[1], INK, dash=True)
    if len(line.series) >= 3:
        _set_line_color(line.series[2], MUTED)
    _minimal_chart_style(line)
    dash.add_chart(line, "B33")

    # --- (3) Line chart: gasto acumulado del mes (Análisis section, left) ---
    acum = LineChart()
    acum.title = "Acumulado del mes"
    acum_cats = Reference(aux, min_col=18, min_row=2, max_row=32)    # R2:R32
    acum_data = Reference(aux, min_col=19, min_row=1, max_row=32)    # S1:S32 (header included)
    acum.add_data(acum_data, titles_from_data=True)
    acum.set_categories(acum_cats)
    acum.height = 9
    acum.width = 12
    if len(acum.series) >= 1:
        _set_line_color(acum.series[0], INK)
    _minimal_chart_style(acum)
    dash.add_chart(acum, "B53")

    # --- (4) Line chart: % ahorro mensual (Análisis section, right) ---
    pct = LineChart()
    pct.title = "% Ahorro mensual"
    pct_cats = Reference(aux, min_col=9, min_row=2, max_row=13)      # I2:I13
    pct_data = Reference(aux, min_col=13, min_row=1, max_row=13)     # M1:M13
    pct.add_data(pct_data, titles_from_data=True)
    pct.set_categories(pct_cats)
    pct.height = 9
    pct.width = 12
    if len(pct.series) >= 1:
        _set_line_color(pct.series[0], INK)
    _minimal_chart_style(pct)
    dash.add_chart(pct, "K53")

    # --- (5) Bar chart: comparativa categorías este mes vs mes anterior ---
    comp = BarChart()
    comp.type = "bar"
    comp.title = "Categorías: este mes vs mes anterior"
    comp_cats = Reference(aux, min_col=21, min_row=2, max_row=11)    # U2:U11
    comp_data = Reference(aux, min_col=22, max_col=23, min_row=1, max_row=11)
    comp.add_data(comp_data, titles_from_data=True)
    comp.set_categories(comp_cats)
    comp.height = 11
    comp.width = 24
    # Style: series 0="Este mes" INK, series 1="Mes anterior" MUTED
    if len(comp.series) >= 1:
        _set_solid_fill(comp.series[0], INK)
    if len(comp.series) >= 2:
        _set_solid_fill(comp.series[1], MUTED)
    _minimal_chart_style(comp)
    dash.add_chart(comp, "B73")


def create_workbook(path: Path) -> None:
    path = Path(path)
    if path.exists():
        raise FileExistsError(
            f"{path} ya existe. Borralo a mano si querés re-crearlo desde cero."
        )

    wb = openpyxl.Workbook()
    # Workbook starts with one default sheet — rename it and add the rest.
    default = wb.active
    default.title = SHEET_ORDER[0]
    for name in SHEET_ORDER[1:]:
        wb.create_sheet(name)

    _build_movimientos(wb["Movimientos"])
    _build_categorias(wb["Categorías"])
    _add_validations(wb["Movimientos"])
    _build_aux(wb["_aux"])
    _build_dashboard(wb["Dashboard"])
    _add_charts(wb)

    wb["_aux"].sheet_state = "hidden"
    wb.active = wb.sheetnames.index("Dashboard")

    # -------------------------------------------------------------------
    # Part D — Cross-cutting: named ranges, freeze panes, gridlines, protection
    # -------------------------------------------------------------------
    dash = wb["Dashboard"]

    # 1. Hide gridlines
    dash.sheet_view.showGridLines = False

    # 2. Freeze panes at row 11 (KPIs stay visible while scrolling charts)
    dash.freeze_panes = "A11"

    # 3. Named ranges
    wb.defined_names["Mes_Seleccionado"] = DefinedName(
        name="Mes_Seleccionado", attr_text="Dashboard!$P$1"
    )
    wb.defined_names["KPI_Ingresos"] = DefinedName(
        name="KPI_Ingresos", attr_text="_aux!$D$1"
    )
    wb.defined_names["KPI_Gastos"] = DefinedName(
        name="KPI_Gastos", attr_text="_aux!$D$2"
    )
    wb.defined_names["KPI_Ahorro"] = DefinedName(
        name="KPI_Ahorro", attr_text="_aux!$D$3"
    )
    wb.defined_names["KPI_PctAhorro"] = DefinedName(
        name="KPI_PctAhorro", attr_text="_aux!$D$4"
    )

    # 4. Sheet protection (P1 stays editable for month selection)
    dash["P1"].protection = Protection(locked=False)
    dash.protection.sheet = True

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Crea gastos.xlsx desde cero.")
    parser.add_argument(
        "--path",
        default="gastos.xlsx",
        help="Ruta del archivo de salida (por defecto: gastos.xlsx en el cwd).",
    )
    args = parser.parse_args()
    create_workbook(Path(args.path))
    print(f"Creado: {args.path}")


if __name__ == "__main__":
    main()
