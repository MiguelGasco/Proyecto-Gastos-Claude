"""Append / edit-last / delete-last rows in tblMov of gastos.xlsx."""
from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font

from scripts.categories import VALID_TYPES, is_valid

# Skill-convention constants (must match init_workbook.py)
_FMT_CURRENCY = '#,##0.00 "€";(#,##0.00 "€");"-"'
_FMT_DATE     = "dd/mm/yyyy"
_COLOR_BLACK  = "FF000000"
_COLOR_BLUE   = "FF0000FF"


def _font_black(cell) -> None:
    cell.font = Font(name="Arial", size=11, color=_COLOR_BLACK)


def _font_blue(cell) -> None:
    cell.font = Font(name="Arial", size=11, color=_COLOR_BLUE)


SHEET = "Movimientos"
FIRST_DATA_ROW = 2
NUM_COLS = 5  # Fecha, Tipo, Categoría, Importe, Descripción


class InvalidMovementError(ValueError):
    """Raised when a movement's fields don't validate."""


def _validate(tipo: str, categoria: str, importe: float) -> None:
    if tipo not in VALID_TYPES:
        raise InvalidMovementError(
            f"tipo inválido: {tipo!r}. Debe ser uno de {VALID_TYPES}"
        )
    if not is_valid(tipo, categoria):
        raise InvalidMovementError(
            f"categoría {categoria!r} no es válida para tipo {tipo!r}"
        )
    if importe <= 0:
        raise InvalidMovementError(
            f"importe debe ser positivo, recibido: {importe}"
        )


def _last_data_row(ws) -> int:
    """Return the row number of the bottom-most filled data row, or 1 if empty."""
    last = 1
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_col=1, values_only=False):
        if row[0].value is None:
            break
        last = row[0].row
    return last


def add_movement(
    path: Path,
    fecha: date,
    tipo: str,
    categoria: str,
    importe: float,
    descripcion: str = "",
) -> None:
    _validate(tipo, categoria, importe)

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]

    last = _last_data_row(ws)
    target_row = last + 1 if last >= FIRST_DATA_ROW else FIRST_DATA_ROW
    c1 = ws.cell(row=target_row, column=1, value=fecha)
    c1.number_format = _FMT_DATE
    _font_black(c1)
    c2 = ws.cell(row=target_row, column=2, value=tipo)
    _font_black(c2)
    c3 = ws.cell(row=target_row, column=3, value=categoria)
    _font_black(c3)
    c4 = ws.cell(row=target_row, column=4, value=float(importe))
    c4.number_format = _FMT_CURRENCY
    _font_blue(c4)  # blue = hardcoded input (user-entered amount)
    c5 = ws.cell(row=target_row, column=5, value=descripcion or "")
    _font_black(c5)

    wb.save(path)


def edit_last(
    path: Path,
    fecha: Optional[date] = None,
    tipo: Optional[str] = None,
    categoria: Optional[str] = None,
    importe: Optional[float] = None,
    descripcion: Optional[str] = None,
) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]
    last = _last_data_row(ws)
    if last < FIRST_DATA_ROW:
        raise InvalidMovementError("la tabla está vacía, nada que editar")

    # Resolve final values (use existing if not overridden) before validating
    current = {
        "fecha": ws.cell(row=last, column=1).value,
        "tipo": ws.cell(row=last, column=2).value,
        "categoria": ws.cell(row=last, column=3).value,
        "importe": ws.cell(row=last, column=4).value,
        "descripcion": ws.cell(row=last, column=5).value,
    }
    final_tipo = tipo if tipo is not None else current["tipo"]
    final_cat = categoria if categoria is not None else current["categoria"]
    final_imp = importe if importe is not None else current["importe"]
    _validate(final_tipo, final_cat, float(final_imp))

    if fecha is not None:
        c = ws.cell(row=last, column=1, value=fecha)
        c.number_format = _FMT_DATE
        _font_black(c)
    if tipo is not None:
        c = ws.cell(row=last, column=2, value=tipo)
        _font_black(c)
    if categoria is not None:
        c = ws.cell(row=last, column=3, value=categoria)
        _font_black(c)
    if importe is not None:
        c = ws.cell(row=last, column=4, value=float(importe))
        c.number_format = _FMT_CURRENCY
        _font_blue(c)
    if descripcion is not None:
        c = ws.cell(row=last, column=5, value=descripcion)
        _font_black(c)

    wb.save(path)


def delete_last(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]
    last = _last_data_row(ws)
    if last < FIRST_DATA_ROW:
        raise InvalidMovementError("la tabla está vacía, nada que borrar")

    for c in range(1, NUM_COLS + 1):
        ws.cell(row=last, column=c).value = None

    wb.save(path)


def _parse_fecha(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Añade, edita o borra el último movimiento en gastos.xlsx."
    )
    parser.add_argument("--path", default="gastos.xlsx")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--edit-last", action="store_true", dest="edit_last_flag")
    mode.add_argument("--delete-last", action="store_true", dest="delete_last_flag")

    parser.add_argument("--fecha", type=_parse_fecha)
    parser.add_argument("--tipo", choices=VALID_TYPES)
    parser.add_argument("--categoria")
    parser.add_argument("--importe", type=float)
    parser.add_argument("--descripcion", default=None)

    args = parser.parse_args()
    path = Path(args.path)

    if args.delete_last_flag:
        delete_last(path)
        print("OK: última fila borrada.")
        return

    if args.edit_last_flag:
        edit_last(
            path,
            fecha=args.fecha,
            tipo=args.tipo,
            categoria=args.categoria,
            importe=args.importe,
            descripcion=args.descripcion,
        )
        print("OK: última fila actualizada.")
        return

    # Append mode — all required fields must be present
    missing = [
        f for f, v in [
            ("--fecha", args.fecha),
            ("--tipo", args.tipo),
            ("--categoria", args.categoria),
            ("--importe", args.importe),
        ] if v is None
    ]
    if missing:
        parser.error(f"faltan flags requeridos para añadir: {', '.join(missing)}")

    add_movement(
        path,
        fecha=args.fecha,
        tipo=args.tipo,
        categoria=args.categoria,
        importe=args.importe,
        descripcion=args.descripcion or "",
    )
    print("OK: movimiento añadido.")


if __name__ == "__main__":
    main()
